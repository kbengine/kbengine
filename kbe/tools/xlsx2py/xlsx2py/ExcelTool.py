# -*- coding: gb2312 -*-
# written by kebiao, 2010/08/20
# update dependency to openpyxl by 1Pixel, 2019/12/19

import openpyxl
import os
import sys

class ExcelTool:
    """
    简单的封装excel各种操作
    系统要求， windows系统， 安装python2.6以及pywin32-214.win32-py2.6.exe, 以及ms office
    """
    def __init__(self, fileName):
        #try:
        #   self.close()
        #except:
        #   pass

        self.__workbook = None

        self.fileName = os.path.abspath(fileName)
        
    def getWorkbookEx(self, auto_create = False):
        try:
            self.__workbook = openpyxl.open(self.fileName)
            return True
        except:
            pass
            
        if auto_create:
            self.__workbook = openpyxl.Workbook()
            return True
            
        return False
        
    def getXLSX(self):
        return self.__workbook


    def close(self, saveChanges = False):
        """
        关闭excel应用
        """
        
        if saveChanges:
            self.__workbook.save(self.fileName)
            
        self.__workbook.close()

    def getSheetCount(self):
        """
        获得工作表个数
        """
        return len(self.__workbook.worksheets)

    def getSheetNameByIndex(self, index):
        """
        获得excel上指定索引位置上的表名称
        """
        return self.__workbook.worksheets[index].title

    def getSheetByIndex(self, index):
        """
        获得excel上指定索引位置上的表
        """
        try:
            return self.__workbook.worksheets[index]
        except:
            return None
    def __getRowCountOnSheet(self, sheet):
        cc = sheet.max_column + 1
        canRun = True
        
        while cc > 0 and canRun:
            cc -= 1
            for i in range(1, sheet.max_row + 1):
                if sheet.cell(i, cc).value is not None:
                    canRun = False
        return cc

    def getRowCount(self, sheetIndex):
        """
        获得一排有多少元素
        """
        ws = self.__workbook.worksheets[sheetIndex]
        return self.__getRowCountOnSheet(ws)
        
        cc = ws.max_column + 1
        canRun = True
        
        while cc > 0 and canRun:
            cc -= 1
            for i in range(1, ws.max_row + 1):
                if ws.cell(i, cc).value is not None:
                    canRun = False
        return cc

    def getColCount(self, sheetIndex):
        """
        获得一列有多少元素
        """
        return self.__workbook.worksheets[sheetIndex].max_row
        
    def getValue(self, sheet, row, col):
        """
        获得某个工作表的某个位置上的值
        """
        return sheet.cell(row, col).value

    def getText(self, sheet, row, col):
        """
        获得某个工作表的某个位置上的值
        """
        return str(sheet.cell(row, col).value)

    def getRowValues(self, sheet, row):
        """
        整排
        """
        cc = self.__getRowCountOnSheet(sheet)
        return [sheet.cell(row+1, i).value for i in range(1, cc + 1)]

    def getSheetRowIters(self, sheet, row):
        """
        行迭代器
        """
        return sheet.Cells(1).CurrentRegion.Rows

    def getSheetColIters(self, sheet, col):
        """
        列迭代器
        """
        return sheet.Cells(1).CurrentRegion.Columns

    def getColValues(self, sheet, col):
        """
        整列
        """
        rc = sheet.max_row
        return [sheet.cell(i, col+1).value for i in range(1, rc+1)]

#---------------------------------------------------------------------
#   使用例子
#---------------------------------------------------------------------
def main():
    xbook = ExcelTool("test.xlsx")

    print("sheetCount=%i" % xbook.getSheetCount())

    for x in range(1, xbook.getSheetCount()):
       print( "      ", xbook.getSheetNameByIndex(x))

    print( "sheet1:rowCount=%i, colCount=%i" % (xbook.getRowCount(1), xbook.getColCount(1)))

    for r in range(1, xbook.getRowCount(1)+1):
        for c in range(1, xbook.getColCount(1)+1):
            val = xbook.getValue(xbook.getSheetByIndex(2), r, c)
            if val:
                print( "DATA:", val)

if __name__ == "__main__":
    main()




